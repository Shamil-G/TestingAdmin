from db_oracle.connect import get_connection
import config as cfg
from openpyxl import load_workbook
import datetime
import os.path


def load_theme(id_task, file_name):
    s_now = datetime.datetime.now()
    if cfg.os == 'unix':
       file_path = cfg.REPORTS_PATH + '/' + file_name
    else:
       file_path = cfg.REPORTS_PATH + '\\' + file_name

    # Нормируем путь к файлу по слэшам
    path = os.path.normpath(file_path)

    print("Загрузка стартовала: " + s_now.strftime("%d-%m-%Y %H:%M:%S") + ' : ' + file_name + ' : ' + file_path)

    if not os.path.isfile(file_path):
        print("File not exists: " + str(os.path.isfile(file_path)))
        return file_name
    print("Load Theme with Excel file: " + str(os.path.isfile(file_path)))

    wb = load_workbook(path)
    print("Книга загружена: " + path)
    sheet = wb.active

    print("Подключаем БД")

    con = get_connection()
    cursor = con.cursor()
    # Создадим новое задание
    id_theme = cursor.callfunc('admin.theme_new', int, (id_task, file_name))
    if not id_theme:
        print('Ошибка регистрации нового задания...')
    id_quest = 0
    id_prev_quest = -1
    order_num = 0
    for i in range(2, sheet.max_row+1):
        id_curr_quest = sheet.cell(row=i, column=2).value
        quest = sheet.cell(row=i, column=3).value
        correctly = sheet.cell(row=i, column=4).value
        answer = sheet.cell(row=i, column=5).value
        order_num = order_num + 1
        if id_curr_quest != id_prev_quest:
            id_quest = id_quest + 1
            order_num = 1
            id_question = cursor.callfunc("admin.add_question", str, [id_theme, id_quest, quest])

        cmd = "insert into answers q (id_answer, id_question, order_num_answer, correctly, answer) " \
              "values ( seq_answer.nextval, " + str(id_question) + ", " + str(order_num) + ", '" + correctly + "', '" + str(answer) + "')"
        print('+++ CMD: ' + cmd)
        cursor.execute(cmd)
        id_prev_quest = id_curr_quest

    con.commit()
    con.close()
    now = datetime.datetime.now()
    print("Загрузка завершена: " + now.strftime("%d-%m-%Y %H:%M:%S"))
    return id_theme


def load_persons(id_task, file_name):
    s_now = datetime.datetime.now()

    if cfg.os == 'unix':
       file_path = cfg.REPORTS_PATH + '/' + file_name
    else:
       file_path = cfg.REPORTS_PATH + '\\' + file_name

    # Нормируем путь к файлу по слэшам
    path = os.path.normpath(file_path)

    print("Загрузка стартовала: " + s_now.strftime("%d-%m-%Y %H:%M:%S") + ' : ' + file_name + ' : ' + file_path)
    if not os.path.isfile(file_path):
        print("File not exists: " + str(os.path.isfile(file_path)))
        return file_name

    print("Load List Persons with Excel file: " + str(os.path.isfile(file_path)))

    wb = load_workbook(path)
    sheet = wb.active
    print("Книга загружена: " + path + ", строк: " + str(sheet.max_row+1))

    print("Подключаем БД")

    con = get_connection()
    cursor = con.cursor()
    for i in range(2, sheet.max_row+1):
        iin = sheet.cell(row=i, column=2).value
        fio = sheet.cell(row=i, column=3).value
        if cfg.debug_level > 2:
            print(str(i) + ". id_task: " + str(id_task) + ", iin: " + str(iin) + ", fio: " + str(fio))
        if not iin or not fio:
            break
        cursor.callproc('admin.load_person', [id_task, iin, fio])
    con.commit()
    con.close()
    now = datetime.datetime.now()
    print("Загрузка завершена: " + now.strftime("%d-%m-%Y %H:%M:%S"))


if __name__ == "__main__":
    print("Тестируем загрузку Excel!")
